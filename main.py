import openpyxl as excel
from datetime import date
from pathlib import Path

def main():

    name, subject, items = get_Invoice_detail()
    path = "Invoice/invoice-template.xlsx"
    invoice_book = excel.load_workbook(path)
    invoice_sheet = invoice_book.active

    invoice_sheet["B4"] = name
    invoice_sheet["C10"] = subject
    invoice_sheet["G3"] = get_date()

    total = 0
    for index, item in enumerate(items):
        summary, quantity, price = item
        row = 15 + index

        subtotal = quantity * price
        total += subtotal

        invoice_sheet.cell(row, 2, summary)
        invoice_sheet.cell(row, 5, quantity)
        invoice_sheet.cell(row, 6, price)
        invoice_sheet.cell(row, 7, subtotal)

    invoice_sheet["C11"] = total

    downloads_path = str(Path.home() / "Downloads")
    file_name = input("保存するファイル名を記入してください: ")

    invoice_book.save(f'{downloads_path}/{file_name}.xlsx')

def get_Invoice_detail():
    name = input("取引先の名前を入力してください： ")
    subject = input("件名を入力してください： ")

    items = []
    is_continue = True
    while is_continue:
        summary = input("概要を記入してください: ")
        quantity = int(input("個数を入力してください:　"))

        if quantity > 500:
            decide = input("入力された'個数'は正しいですか？: y/n")
            if decide == 'n' or decide == 'ｎ':
                quantity = int(input("個数を入力してください:　"))

        price = int(input("金額を入力してください: "))

        item = [summary, quantity, price]
        items.append(item)

        decide = input("入力を続けますか？ y/n: ")

        if decide == 'n' or decide == 'ｎ':
            is_continue = False

    return name, subject, items

def get_date():
    today = date.today()
    return today.strftime("%d/%m/%Y")

if __name__ == "__main__":
    main()